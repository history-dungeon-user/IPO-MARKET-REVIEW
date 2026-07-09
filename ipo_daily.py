# -*- coding: utf-8 -*-
"""
ipo_daily.py — IPO 통합 일일 자동화 (예심 + 공모~상장 + 3탭 엑셀 + 웹페이지)
================================================================================
실행: python ipo_daily.py          (kind_weekly.py 와 같은 폴더에 두고 실행)
필요: pip install requests beautifulsoup4 openpyxl

구성
  [1] KIND 예심 수집·3표          → kind_weekly.py 로직 재사용 (import)
  [2] KIND 신규상장 페이지 수집    → 상장일 확정 소스 (listingcompany.do)
  [3] DART 신고서 추적·발췌        → 승인법인 공모 데이터 (OpenDART API)
  [4] 3탭 엑셀 생성                → IPO_analysis.xlsx
  [5] index.html 재생성            → kind_weekly.py 가 담당

검증된 발췌 규칙 (2026-07-07 확정)
  - 최신 정정본 우선: 접수일 최신 rcept_no 1건만 사용
  - 기재정정/발행조건확정 문서 내 정정 전·후 병존 → "정정 후" 값만 취함
  - 공모개요표: 증권수량/모집(매출)가액(=밴드하단)/모집(매출)총액
  - 인수대가 + 주석 수수료율(%) / 성과수수료 별도
  - 밸류에이션: 평가모형(PER/PSR/EV/EBITDA) + 비교회사 배수 (표기 변형 허용)
  - 수요예측 참여내역표: 건수·경쟁률 행의 '합계' 열
  - 실적보고서: (일반)청약경쟁률 = 일반투자자 청약수량 ÷ 최종배정수량
  - 상장일: KIND 신규상장 페이지 등재 시 확정 (그 전엔 상장예정일)
"""
import os, re, sys, json, zipfile, io, datetime
import requests

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)

# ── 설정 ─────────────────────────────────────────────────────────────
DART_API_KEY = os.environ.get("DART_API_KEY", "")   # 환경변수 또는 아래에 직접 입력
if not DART_API_KEY:
    # D:\dart-mcp\dart_mcp.py 에서 키 자동 탐색 시도
    try:
        sys.path.insert(0, r"D:\dart-mcp")
        import dart_mcp as _dm
        for attr in ("DART_API_KEY", "API_KEY", "DART_KEY"):
            if hasattr(_dm, attr):
                DART_API_KEY = getattr(_dm, attr); break
    except Exception:
        pass

OFFERING_JSON = os.path.join(BASE, "offering_master.json")   # 공모 추적 정본
LISTING_JSON  = os.path.join(BASE, "listing_master.json")    # 신규상장 정본
XLSX_OUT      = os.path.join(BASE, "IPO_analysis.xlsx")

DART = "https://opendart.fss.or.kr/api"
KIND = "https://kind.krx.co.kr"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/126.0 Safari/537.36",
    "Referer": KIND + "/listinvstg/listingcompany.do",
}

# 승인법인 corp_code 시드 (검증 완료 · 신규 승인법인은 자동 매칭 후 추가)
CORP_SEED = {
 "기도산업":"00229085","브릴스":"01801026","니어스랩":"01588897","스카이랩스":"01586923",
 "해치텍":"01947443","와이즈플래닛컴퍼니":"01867688","케이앤에스아이앤씨":"00920731",
 "빅웨이브로보틱스":"01722066","인제니아테라퓨틱스":"02031369","에이치엘지노믹스":"00624244",
 "레메디":"01501764","딜리셔스":"01461730",
}

TODAY = datetime.date.today().strftime("%Y%m%d")


# ════════════════════════════════════════════════════════════════════
# [2] KIND 신규상장 페이지 (상장일 확정 소스)
# ════════════════════════════════════════════════════════════════════
def fetch_new_listings(sess, year=None):
    """listingcompany.do 신규상장 목록 → [{회사명, 상장일, 상장유형, 시장}]
    ※ F12 실측(2026-07-08) 확정 파라미터:
       method=searchListingTypeSub, forward=listingtype_sub,
       listTypeArrStr=01|02|03|04|05|  (신규·이전·재상장·폐지·합병)
       secuGrpArrStr=0|ST|FS|MF|SC|RT|IF|DR|  (전체 증권구분)
    ※ Main(폼)이 아니라 Sub(데이터)를 호출해야 목록이 반환됨."""
    year = year or datetime.date.today().year
    url = KIND + "/listinvstg/listingcompany.do"
    # 폼 세션 확보 (Referer/쿠키)
    try:
        sess.get(url + "?method=searchListingTypeMain", headers=HEADERS, timeout=20)
    except Exception:
        pass
    data = {
        "method": "searchListingTypeSub",
        "forward": "listingtype_sub",
        "currentPageSize": "500", "pageIndex": "1",
        "orderMode": "1", "orderStat": "D",
        "listTypeArrStr": "01|02|03|04|05|",
        "secuGrpArrStr": "0|ST|FS|MF|SC|RT|IF|DR|",
        "fromDate": f"{year}-01-01",
        "toDate": datetime.date.today().strftime("%Y-%m-%d"),
    }
    hdr = dict(HEADERS)
    hdr["Content-Type"] = "application/x-www-form-urlencoded; charset=UTF-8"
    hdr["Origin"] = KIND
    hdr["Referer"] = url + "?method=searchListingTypeMain"
    try:
        r = sess.post(url, data=data, headers=hdr, timeout=30)
        r.raise_for_status()
    except Exception as e:
        print(f"  [경고] 신규상장 페이지 접속 실패: {e}")
        return None
    if "잠시 후" in r.text or len(r.text) < 500:
        print("  [경고] 신규상장 응답이 비정상(차단 또는 빈 응답)")
        return None
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(r.text, "html.parser")
    rows = []
    for tr in soup.select("table tbody tr, table tr"):
        tds = tr.find_all("td")
        if len(tds) < 4:
            continue
        name = tds[0].get_text(strip=True)
        date = tds[1].get_text(strip=True).replace(".", "-")
        ltype = tds[2].get_text(strip=True) if len(tds) > 2 else ""
        # 시장 판별: 회사명 셀의 아이콘(alt/src/class) 전체를 훑음
        market = ""
        img = tds[0].find("img")
        blob = ""
        if img:
            blob = " ".join([img.get("alt", ""), img.get("src", ""),
                             " ".join(img.get("class", []))]).lower()
        # 행 전체 텍스트/클래스도 보조 신호로 활용
        row_txt = tr.get_text(" ", strip=True)
        if "유가" in blob or "market_yu" in blob or "icn_t_yu" in blob or "유가" in row_txt[:0]:
            market = "유가"
        elif "코넥스" in blob or "knx" in blob or "icn_t_kn" in blob or "kn.gif" in blob:
            market = "코넥스"
        elif "코스닥" in blob or "market_ko" in blob or "icn_t_ko" in blob or "ko.gif" in blob:
            market = "코스닥"
        # 코넥스는 IPO(코스닥·유가) 대상이 아니므로 제외
        if market == "코넥스":
            continue
        if re.match(r"\d{4}-\d{2}-\d{2}", date) and name:
            rows.append({"회사명": name, "상장일": date,
                         "상장유형": ltype, "시장": market})
    return rows


# ════════════════════════════════════════════════════════════════════
# [3] DART 신고서 추적·발췌
# ════════════════════════════════════════════════════════════════════
def dart_list(corp_code, bgn="20260101"):
    """공시목록 (발행공시 C)"""
    r = requests.get(f"{DART}/list.json", params={
        "crtfc_key": DART_API_KEY, "corp_code": corp_code,
        "bgn_de": bgn, "end_de": TODAY, "pblntf_ty": "C",
        "page_count": 100}, timeout=30)
    j = r.json()
    return j.get("list", []) if j.get("status") == "000" else []

def dart_document_text(rcept_no):
    """원문 ZIP → 전체 텍스트 (표는 '|' 구분으로 평탄화)"""
    r = requests.get(f"{DART}/document.xml", params={
        "crtfc_key": DART_API_KEY, "rcept_no": rcept_no}, timeout=60)
    zf = zipfile.ZipFile(io.BytesIO(r.content))
    texts = []
    for name in zf.namelist():
        raw = zf.read(name)
        for enc in ("utf-8", "euc-kr", "cp949"):
            try:
                s = raw.decode(enc); break
            except Exception:
                s = raw.decode("utf-8", "ignore")
        # 표 셀 경계 유지하며 태그 제거
        s = re.sub(r"</T[DHU][^>]*>", " | ", s, flags=re.I)
        s = re.sub(r"</TR[^>]*>", "\n", s, flags=re.I)
        s = re.sub(r"<[^>]+>", " ", s)
        s = re.sub(r"&[a-z]+;", " ", s)
        s = re.sub(r"[ \t]+", " ", s)
        texts.append(s)
    return "\n".join(texts)

def latest_filing(filings):
    """최신 증권신고서 1건 + 상태 판별 + 실적보고서 여부"""
    reg, perf = None, None
    for f in sorted(filings, key=lambda x: x.get("rcept_dt", ""), reverse=True):
        nm = f.get("report_nm", "")
        if "증권발행실적보고서" in nm and perf is None:
            perf = f
        if "증권신고서" in nm and "실적" not in nm and reg is None:
            reg = f
    status = "미제출"
    if reg:
        nm = reg["report_nm"]
        if "발행조건확정" in nm: status = "발행조건확정"
        elif "기재정정" in nm or "정정" in nm: status = "정정"
        else: status = "최초"
    return reg, perf, status

# ── 발췌기 (검증 규칙 구현) ──────────────────────────────────────────
def _after_correction(text):
    """'정정 후' 마지막 구간 우선 — 없으면 전체"""
    parts = re.split(r"정정\s*후", text)
    return parts[-1] if len(parts) > 1 else text

def _num(s):
    try: return float(str(s).replace(",", "").strip())
    except Exception: return None

def extract_registration(text):
    """증권신고서(최초/정정/확정 공통) 발췌"""
    out = {}
    t = _after_correction(text)

    m = re.search(r"희망공모가액[^\d]{0,20}([\d,]+)\s*원?\s*[~∼]\s*([\d,]+)\s*원", t) or \
        re.search(r"희망공모가액[^\d]{0,20}([\d,]+)\s*원?\s*[~∼]\s*([\d,]+)\s*원", text)
    if m: out["밴드"] = f"{m.group(1)}~{m.group(2)}"

    m = re.search(r"보통주\s*\|\s*([\d,]+)\s*\|\s*[\d,]+\s*\|\s*([\d,]+)\s*\|\s*([\d,]+)\s*\|", t)
    if m:
        out["공모주식수"] = m.group(1)
        out["밴드하단가"] = m.group(2)
        out["공모금액"] = m.group(3)

    m = re.search(r"인수대가[^\n]{0,200}?([\d.]+)\s*%", t)
    if m: out["수수료율"] = m.group(1) + "%"
    m = re.search(r"대표주관회사\s*\|\s*([가-힣A-Za-z()㈜ ]+?)\s*\|[^\n]*?\|\s*([\d,]+)\s*\|\s*총액인수", t)
    if m: out["대표주관"], out["인수대가"] = m.group(1).strip(), m.group(2)
    co = re.findall(r"공동주관회사\s*\|\s*([가-힣A-Za-z()㈜ ]+?)\s*\|", t)
    if co: out["공동주관"] = ", ".join(x.strip() for x in co)

    m = re.search(r"(비교회사|유사회사|적용)\s*(P/?E R?|PER|PSR|EV/EBITDA)[^\d]{0,30}([\d.]+)\s*배?", t, re.I)
    if m: out["멀티플"] = f"{m.group(3)} ({m.group(2).upper().replace(' ','')})"

    m = re.search(r"수요예측\s*일시?\s*\|?\s*(\d{4})[.\s년]+(\d{1,2})[.\s월]+(\d{1,2})"
                  r"[일)월화수목금토(\s]*[~∼]\s*(?:(\d{4})[.\s년]+)?(\d{1,2})[.\s월]+(\d{1,2})", t)
    if m:
        out["수요예측기간"] = f"{int(m.group(2)):02d}-{int(m.group(3)):02d}~{int(m.group(5)):02d}-{int(m.group(6)):02d}"

    m = re.search(r"청약기[일간][\s\S]{0,160}?(\d{4})[.\s년]+(\d{1,2})[.\s월]+(\d{1,2})"
                  r"[일)월화수목금토(\s]*[~∼]\s*(?:(\d{4})[.\s년]+)?(\d{1,2})[.\s월]+(\d{1,2})", t)
    if m: out["청약일정"] = f"{int(m.group(2)):02d}-{int(m.group(3)):02d}~{int(m.group(5)):02d}-{int(m.group(6)):02d}"

    # 확정본 전용
    m = re.search(r"확정공모가액[^\d]{0,20}([\d,]+)\s*원", t)
    if m: out["확정공모가"] = m.group(1)
    # 참여내역 합계: 건수/경쟁률 행의 마지막 숫자
    seg = re.search(r"수요예측\s*참여\s*내역[\s\S]{0,2500}", t)
    if seg:
        s = seg.group(0)
        m = re.search(r"건수\s*\|([^\n]+)", s)
        if m:
            nums = re.findall(r"[\d,]+", m.group(1))
            if nums: out["참여기관수"] = nums[-1]
        m = re.search(r"경쟁률\s*\|([^\n]+)", s)
        if m:
            nums = re.findall(r"[\d,]+\.\d+|[\d,]+", m.group(1))
            if nums: out["수요예측경쟁률"] = nums[-1]
    return out

def extract_performance(text):
    """증권발행실적보고서: (일반)청약경쟁률 + 상장예정일"""
    out = {}
    seg = re.search(r"청약\s*및\s*배정현황[\s\S]{0,1500}", text)
    if seg:
        m = re.search(r"일반투자자\s*\|([^\n]+)", seg.group(0))
        if m:
            nums = [_num(x) for x in re.findall(r"[\d,]+", m.group(1))]
            nums = [n for n in nums if n]
            # 일반투자자 행: 최초배정수량, 비율, 건수, 청약수량, 금액, 비율, 건수, 최종수량 ...
            if len(nums) >= 8:
                chung, bae = nums[3], nums[7]
                if bae: out["청약경쟁률"] = f"{chung/bae:,.2f}"
    m = re.search(r"상장\s*예정일[^\d]{0,20}(\d{4})[.\s년]+(\d{1,2})[.\s월]+(\d{1,2})", text)
    if m: out["상장예정일"] = f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return out


# ── 강건 발췌기 v2 (검증: 상장완료 12딜 · 인수인 4딜 실측 일치) ──────────
_FIRM_ABBR = [("미래에셋","미래"),("한국투자","한국"),("엔에이치","NH"),("NH투자","NH"),
    ("케이비","KB"),("KB","KB"),("삼성","삼성"),("신한투자","신한"),("대신","대신"),
    ("하나","하나"),("키움","키움"),("유진투자","유진"),("메리츠","메리츠"),
    ("교보","교보"),("현대차","현대차"),("유안타","유안타"),("IBK","IBK"),("BNK","BNK")]
def _abbr_firm(name):
    for k, v in _FIRM_ABBR:
        if k in name: return v
    return re.sub(r"투자증권|증권", "", name or "").strip()
_ROLE_MAP = {"공동대표주관회사":"공동대표","대표주관회사":"대표","대표":"대표",
    "공동주관회사":"공동주관","인수회사":"인수단","인수단":"인수단"}

def extract_demand(t):
    """수요예측 총경쟁률·참여건수(수요예측 참여내역 표 합계 열)"""
    out = {}; best = None
    for m in re.finditer(r"경쟁률[^\n]{0,8}?\|([^\n]+)", t):
        vals = []
        for x in re.findall(r"[\d,]+\.\d+|[\d,]+", m.group(1)):
            try: vals.append(float(x.replace(",", "")))
            except Exception: pass
        if len(vals) >= 6 and (best is None or vals[-1] > best[0]):
            best = (vals[-1], m.start())
    if best:
        out["수요예측경쟁률"] = round(best[0], 2)
        pre = t[max(0, best[1]-1600):best[1]]; cm = None
        for m in re.finditer(r"건수\s*\|([^\n]+)", pre): cm = m
        if cm:
            cn = re.findall(r"[\d,]+", cm.group(1))
            if cn: out["참여기관수"] = int(cn[-1].replace(",", ""))
    return out

def extract_multiple(t):
    """공모가 산정 요약표: 평가방법 · 평가모형(PER/PBR/EV·EBITDA) 배수"""
    method = None
    m = re.search(r"평가방법\s*\|\s*([가-힣]+)", t)
    if m: method = m.group(1)
    m = re.search(r"②[^\n]{0,4}\|?\s*(?:유사기업|유사회사|비교회사|비교기업|비교대상회사|비교대상기업)"
                  r"\s*(?:평균\s*)?(PER|PBR|EV\s*/?\s*EBITDA)\s*\|?\s*(?:배\s*\|?\s*)?([\d,]+\.?\d*)", t)
    if not m:
        m = re.search(r"(?:유사기업|유사회사|비교회사|비교기업|비교대상회사|비교대상기업)"
                      r"\s*(?:평균\s*)?(PER|PBR|EV\s*/?\s*EBITDA)\s*\|?\s*(?:배\s*\|?\s*)?([\d,]+\.?\d*)\s*(?:배|x|X)", t)
    if m:
        kind = re.sub(r"\s", "", m.group(1)); mult = m.group(2)
        return f"{method + ' · ' if method else ''}{kind} {mult}배"
    return None

def extract_payment_date(t):
    """공모개요 납입기일 (청약기일 다음 마지막 일자)"""
    seg = re.search(r"납\s*입\s*기\s*일[\s\S]{0,260}", t)
    if not seg: return None
    ds = re.findall(r"(\d{4})[.\s년]+(\d{1,2})[.\s월]+(\d{1,2})", seg.group(0))
    if ds:
        y, mo, da = ds[-1]
        return f"{y}-{int(mo):02d}-{int(da):02d}"
    return None

def extract_perf_fee(t):
    """성과수수료 조항 발췌(요약)"""
    m = re.search(r"([^.。\n]{0,70}?([\d.]+)\s*%[^.。\n]{0,45}?성과수수료[^.。\n]{0,70})", t)
    if m: return re.sub(r"\s+", " ", m.group(1)).strip()[:140]
    m = re.search(r"(성과수수료[^.。\n]{0,140})", t)
    if m: return re.sub(r"\s+", " ", m.group(1)).strip()[:140]
    return "없음"

def extract_underwriters(t):
    """인수인 표 → [{증권사,역할,인수금액(억),인수대가(억)}] · 증권사별 정정후(마지막) 유지"""
    rows = []
    for m in re.finditer(
        r"(공동대표주관회사|대표주관회사|공동주관회사|인수회사|인수단|대표)\s*\|\s*"
        r"([가-힣A-Za-z]+증권)\s*\|\s*기명식[^|]*\|\s*([\d,]+)\s*\|\s*([\d,]+)\s*\|\s*([\d,]+)\s*\|\s*총액인수", t):
        role, firm, qty, amt, fee = m.groups()
        rows.append({"증권사": _abbr_firm(firm), "역할": _ROLE_MAP.get(role, role),
            "인수금액": round(int(amt.replace(",", "")) / 1e8, 2),
            "인수대가": round(int(fee.replace(",", "")) / 1e8, 3)})
    last = {}
    for r in rows: last[r["증권사"]] = r   # 정정 전/후 중복 → 마지막(확정) 유지
    return list(last.values())


def enrich_listed_from_dart(listings, corp_map):
    """신규/미백필 상장딜 → 발행조건확정 신고서 자동 발췌 → listing_seed.json 갱신
       + 인수인 표 → underwriter_ledger.json 자동 확장(납입일 기준 누적).
       DART 키 없으면 조용히 통과. 절대 파이프라인을 중단시키지 않음."""
    if not DART_API_KEY:
        print("[ENRICH] DART 키 없음 → 신고서 자동발췌 생략"); return
    seed_path = os.path.join(BASE, "listing_seed.json")
    seed = {x["기업명"]: x for x in _load_listing_seed()}
    changed = 0
    for l in listings:
        nm = l["회사명"]
        ty = (l.get("상장유형") or "")
        if "합병" in ty:            # 스팩합병 상장은 공모개념 없음
            continue
        rec = seed.get(nm)
        if rec is None:
            rec = {"기업명": nm, "진행상태": "상장완료"}; seed[nm] = rec
        # 이미 백필된 딜은 건너뜀(수동검증 우선)
        if str(rec.get("수요예측경쟁률") or "") not in ("", "수집예정"):
            continue
        code = corp_map.get(nm)
        if not code:
            continue
        try:
            reg, perf, status = latest_filing(dart_list(code))
            if not reg:
                continue
            text = dart_document_text(reg["rcept_no"])
        except Exception as e:
            print(f"  [ENRICH 경고] {nm}: {e}"); continue
        info = extract_registration(text)          # 밴드·확정가·대표주관 등
        info.update(extract_demand(text))          # 수요예측경쟁률·참여기관수
        pay = extract_payment_date(text)
        if pay: info["납입일"] = pay
        mult = extract_multiple(text)
        if mult: info["멀티플"] = mult
        fee = extract_perf_fee(text)
        if fee and fee != "없음": info["성과수수료"] = fee
        for k, v in info.items():
            if v not in (None, "", "수집예정"):
                rec[k] = v
        rec["_자동발췌"] = f"{status}·{reg['rcept_no']}"
        changed += 1
        print(f"  [ENRICH] {nm}: 수요예측 {info.get('수요예측경쟁률')} · "
              f"멀티플 {info.get('멀티플')} · 납입 {info.get('납입일')}")
        # 인수인 → 원장 자동 확장
        try:
            _extend_ledger(nm, l, extract_underwriters(text))
        except Exception as e:
            print(f"  [ENRICH 원장경고] {nm}: {e}")
    if changed:
        json.dump(list(seed.values()), open(seed_path, "w", encoding="utf-8"),
                  ensure_ascii=False, indent=1)
        print(f"[ENRICH] listing_seed.json 자동갱신 {changed}건")


def _extend_ledger(company, listing, underwriters):
    """인수인 표를 underwriter_ledger.json 딜원장에 추가하고 집계 재계산.
       청약수수료=기관배정액(75% 가정)×1%(스팩 제외) · 인수수수료=인수대가."""
    lp = os.path.join(BASE, "underwriter_ledger.json")
    if not (os.path.exists(lp) and underwriters):
        return
    L = json.load(open(lp, encoding="utf-8"))
    dl = L.setdefault("딜원장", [])
    if any(d.get("업체") == company for d in dl):
        return
    is_spac = bool(re.search(r"스팩|스펙|기업인수목적", company))
    total = round(sum(u["인수금액"] for u in underwriters), 2)
    ui = []
    for u in underwriters:
        share = (u["인수금액"] / total) if total else 0
        chung = 0.0 if is_spac else round(total * 0.75 * 0.01 * share, 3)
        ui.append({"증권사": u["증권사"], "역할": u["역할"],
                   "인수금액": u["인수금액"], "인수수수료": u["인수대가"],
                   "청약수수료": chung, "전체수수료": round(u["인수대가"] + chung, 3)})
    dl.append({"업체": company, "상장일": listing.get("상장일", ""),
               "시장": listing.get("시장", "코스닥"), "발행금액": total, "인수인": ui})
    # 집계 재계산
    agg = {}
    for d in dl:
        for u in d["인수인"]:
            a = agg.setdefault(u["증권사"], {"증권사": u["증권사"], "인수금액": 0.0,
                "인수수수료": 0.0, "청약수수료": 0.0, "전체수수료": 0.0, "건수": 0})
            a["인수금액"] += u.get("인수금액", 0) or 0
            a["인수수수료"] += u.get("인수수수료", 0) or 0
            a["청약수수료"] += u.get("청약수수료", 0) or 0
            a["전체수수료"] += u.get("전체수수료", 0) or 0
        for u in d["인수인"]:
            if u["역할"] in ("대표", "공동대표"):
                agg[u["증권사"]]["건수"] += 1
    for a in agg.values():
        for k in ("인수금액", "인수수수료", "청약수수료", "전체수수료"):
            a[k] = round(a[k], 3)
    L["주관사집계"] = sorted(agg.values(), key=lambda r: -r["전체수수료"])
    json.dump(L, open(lp, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(f"  [원장] {company} 추가 → 딜 {len(dl)}건")


def track_offerings(approved_names):
    """승인법인 → corp_code → 최신 신고서 → 발췌 → offering_master 갱신"""
    master = {}
    if os.path.exists(OFFERING_JSON):
        master = json.load(open(OFFERING_JSON, encoding="utf-8"))
    corp_map = dict(CORP_SEED)
    cm_path = os.path.join(BASE, "corp_map.json")
    if os.path.exists(cm_path):
        try: corp_map.update(json.load(open(cm_path, encoding="utf-8")).get("corp_map", {}))
        except Exception: pass

    changes = []
    for name in approved_names:
        code = corp_map.get(name)
        rec = master.get(name, {"회사명": name})
        if not code:
            rec["상태"] = "corp_code 미확보(수동확인)"
            master[name] = rec; continue
        try:
            filings = dart_list(code)
        except Exception as e:
            print(f"  [경고] {name} DART 조회 실패: {e}"); continue
        reg, perf, status = latest_filing(filings)
        new_rcept = reg["rcept_no"] if reg else None
        if new_rcept and rec.get("최신접수번호") != new_rcept:
            try:
                text = dart_document_text(new_rcept)
                rec.update(extract_registration(text))
                changes.append(f"{name}: 신고서 갱신({status}) {new_rcept}")
            except Exception as e:
                print(f"  [경고] {name} 발췌 실패: {e}")
            rec["최신접수번호"] = new_rcept
        if perf and rec.get("실적보고서") != perf["rcept_no"]:
            try:
                text = dart_document_text(perf["rcept_no"])
                rec.update(extract_performance(text))
                changes.append(f"{name}: 실적보고서 반영 {perf['rcept_no']}")
            except Exception as e:
                print(f"  [경고] {name} 실적 발췌 실패: {e}")
            rec["실적보고서"] = perf["rcept_no"]
        rec["상태"] = status if not perf else "청약완료·상장대기"
        rec["corp_code"] = code
        master[name] = rec

    json.dump(master, open(OFFERING_JSON, "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)
    return master, changes


# ════════════════════════════════════════════════════════════════════
# [1] KIND 예심 수집 (kind_weekly 재사용 · 무인 실행용, input() 없음)
# ════════════════════════════════════════════════════════════════════
def collect_kind():
    import kind_weekly as kw
    import datetime as dt
    sess = requests.Session()
    try: sess.get(kw.REFERER, headers=kw.HEADERS, timeout=15)
    except Exception: pass
    fresh = kw.fetch_list(sess)
    master = kw.load_json(kw.MASTER, {"records": {}})
    store = master["records"]
    now = dt.datetime.now().isoformat(timespec="seconds")
    added, changed = [], []
    for rec in fresh:
        k = f"{rec['회사명']}||{rec['청구일']}"
        cur = "심사중" if kw.classify(rec["심사결과"]) == "신청" else "결과확정"
        if k not in store:
            store[k] = {**rec, "상태": cur, "최초수집": now, "최근확인": now,
                        "상태변경이력": [{"상태": cur, "시각": now}]}
            added.append(f"신규\t{rec['청구일']}\t{rec['회사명']}\t{rec['심사결과']}")
        else:
            it = store[k]; it["최근확인"] = now
            for fld in ("시장", "bizProcNo"):
                if rec.get(fld) and not it.get(fld): it[fld] = rec[fld]
            if it.get("상태") != cur or kw.nz(it.get("심사결과","")) != kw.nz(rec["심사결과"]):
                it["상태변경이력"] = it.get("상태변경이력", []) + [
                    {"상태": cur, "시각": now, "심사결과": rec["심사결과"]}]
                it["상태"], it["결과확정일"], it["심사결과"] = cur, rec["결과확정일"], rec["심사결과"]
                changed.append(f"상태변경\t{rec['청구일']}\t{rec['회사명']}\t-> {rec['심사결과']}")
    kw.save_json(kw.MASTER, master)

    fin = kw.load_json(kw.FINDATA, {})
    for k, v in kw.SEED_FIN.items():
        if k not in fin: fin[k] = {**v, "출처": "시드(검증완료)"}

    def result_date(rec):
        d = (rec.get("결과확정일") or "").strip()
        return d or kw.DATE_OVERRIDE.get(kw.fkey(rec["회사명"], rec["청구일"]), "")

    t_apply, t_appr, t_drop, need = [], [], [], []
    for rec in fresh:
        c = kw.classify(rec["심사결과"])
        if rec["상장유형"].replace(" ", "") == "재상장" and c in ("승인", "철회미승인"):
            continue
        if c == "신청": t_apply.append(rec)
        elif c == "승인" and result_date(rec).startswith(kw.YEAR): t_appr.append(rec)
        elif c == "철회미승인" and result_date(rec).startswith(kw.YEAR): t_drop.append(rec)
        else: continue
        if kw.fkey(rec["회사명"], rec["청구일"]) not in fin and rec.get("bizProcNo"):
            need.append(rec)
    for rec in need:
        try:
            d = kw.fetch_detail(sess, rec["bizProcNo"])
            fin[kw.fkey(rec["회사명"], rec["청구일"])] = {
                "매출액": d["매출액"], "순이익": d["순이익"], "자기자본": d["자기자본"],
                "업종": d["업종"], "단위": d["단위"],
                "유가": "Y" if rec.get("시장") == "유가" else "",
                "출처": f"KIND상세({TODAY})"}
        except Exception as e:
            print(f"  [경고] 상세 실패 {rec['회사명']}: {e}")
    kw.save_json(kw.FINDATA, fin)

    def build(rows, datefn, status=None):
        out = []
        for rec in rows:
            f = fin.get(kw.fkey(rec["회사명"], rec["청구일"]), {})
            unit = f.get("단위", "백만원"); cy = rec["청구일"][:4]
            base_year = "2024" if "FY24" in unit else (str(int(cy)-1) if cy.isdigit() else "")
            out.append({"일자": datefn(rec), "회사명": rec["회사명"],
                "유가": f.get("유가") or ("Y" if rec.get("시장")=="유가" else ""),
                "상장유형": rec["상장유형"], "심사결과": rec["심사결과"],
                "기준연도": base_year, "매출액": f.get("매출액",""),
                "순이익": f.get("순이익",""), "자기자본": f.get("자기자본",""),
                "업종": f.get("업종",""), "단위": unit,
                "대표주관사": kw.abbr(rec["상장주선인"]),
                "현황": status if status is not None else (rec.get("심사결과") or "")})
        out.sort(key=lambda r: r["일자"], reverse=True)
        return out

    # ── 승인 탭: live 목록이 아닌 누적 master 기준 (올해 승인 전체·상장완료 포함) ──
    #   승인 = 심사승인/상장승인/공모철회/상장철회 (재상장 제외) · 일자 = 승인일(결과확정일)
    APPMAP = {"심사승인": "상장전", "상장승인": "상장완료",
              "공모철회": "공모철회", "상장철회": "상장철회"}
    lseed = {_norm_name(x["기업명"]): (x.get("상장일") or "")
             for x in _load_listing_seed()}
    appr = []
    for rec in store.values():
        res = (rec.get("심사결과") or "").replace(" ", "")
        ty = (rec.get("상장유형") or "").replace(" ", "")
        rd = (rec.get("결과확정일") or "")
        if res not in APPMAP or ty == "재상장" or not rd.startswith(kw.YEAR):
            continue
        f = fin.get(kw.fkey(rec.get("회사명",""), rec.get("청구일","")), {})
        unit = f.get("단위", "백만원"); cy = (rec.get("청구일") or "")[:4]
        base_year = "2024" if "FY24" in unit else (str(int(cy)-1) if cy.isdigit() else "")
        ld = lseed.get(_norm_name(rec.get("회사명","")), "") if res == "상장승인" else ""
        hyun = APPMAP[res] + (f" ({ld})" if ld and ld != "수집예정" else "")
        appr.append({"일자": rd, "회사명": rec.get("회사명",""),
            "유가": f.get("유가") or ("Y" if rec.get("시장")=="유가" else ""),
            "상장유형": rec.get("상장유형",""), "심사결과": rec.get("심사결과",""),
            "기준연도": base_year, "매출액": f.get("매출액",""),
            "순이익": f.get("순이익",""), "자기자본": f.get("자기자본",""),
            "업종": f.get("업종",""), "단위": unit,
            "대표주관사": kw.abbr(rec.get("상장주선인","")),
            "현황": hyun})
    appr.sort(key=lambda r: r["일자"], reverse=True)

    DATA = {"asof": kw.TODAY, "year": kw.YEAR,
            "fin_basis": "청구일 직전 사업연도 · 별도/개별 재무제표 · 단위 백만원(별도 표기 시 USD)",
            "tables": {"신청": build(t_apply, lambda r: r["청구일"], "예심 진행중"),
                       "승인": appr,
                       "철회미승인": build(t_drop, result_date)}}
    DATA["counts"] = {k: len(v) for k, v in DATA["tables"].items()}
    html = kw.TEMPLATE.replace("__DATA__", json.dumps(DATA, ensure_ascii=False))
    open(kw.SITE, "w", encoding="utf-8").write(html)
    print(f"[KIND] 예심 갱신: 신규 {len(added)} · 변경 {len(changed)} · "
          f"표 {DATA['counts']}")
    return DATA, master


# ════════════════════════════════════════════════════════════════════
# [4] 3탭 엑셀 생성
# ════════════════════════════════════════════════════════════════════
V = "수집예정"; NA = "해당없음"

def _load_listing_seed():
    p = os.path.join(BASE, "listing_seed.json")
    return json.load(open(p, encoding="utf-8")) if os.path.exists(p) else []

def _norm_name(s):
    """상장사 이름 매칭용 정규화 키.
    시드 표기와 KIND 라이브 표기의 철자 차이를 흡수한다.
    예) 신한스팩18호 = 신한제18호스팩, 교보스팩20호 = 교보20호스팩,
        키움히어로스펙2호 = 키움히어로제2호스팩, NH스팩33호 = 엔에이치스팩33호,
        케이뱅크 (유가) = 케이뱅크
    표시용 이름(기업명)은 건드리지 않고, 매칭 키로만 사용한다.
    """
    s = (s or "").lower().strip()
    s = re.sub(r"\(.*?\)", "", s)      # (유가)/(코스닥) 등 시장 표기 제거
    s = re.sub(r"\s+", "", s)          # 공백 제거
    s = s.replace("엔에이치", "nh")     # 스폰서 약어 통일
    if "스팩" in s or "스펙" in s:      # 스팩류만 강한 정규화(제/호/스팩 순서 차이 흡수)
        s = s.replace("스팩", "").replace("스펙", "")
        s = s.replace("제", "").replace("호", "")
    return s

def build_excel(kind_data, kind_master, offering, listings):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    FONT = "맑은 고딕"
    def F(bold=False, size=10, color="000000", italic=False):
        return Font(name=FONT, bold=bold, size=size, color=color, italic=italic)
    CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIG = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
    def fill(h): return PatternFill("solid", fgColor=h)
    HDR = fill("1F3864"); SUB = fill("D9E1F2")
    C_GRAY="808080"; C_BLUE="2E5B9F"; C_GREEN="2E7D4F"; C_ORANGE="C55A11"; C_RED="9A3B3B"
    NUMCOLS = {4, 5, 16}

    wb = Workbook()

    # ── 탭1 예심 ──
    ws1 = wb.active; ws1.title = "1_예심"
    r = 1
    ws1.cell(r,1,"IPO 상장예비심사 현황").font = F(True,14); r += 1
    ws1.cell(r,1,f"기준일 {kind_data['asof']} · {kind_data['fin_basis']}").font = F(size=9,color="808080"); r += 2
    cols1 = ["일자","회사명","시장","상장유형","기준연도","매출액","순이익","자기자본","업종","대표주관사","현황"]
    for block, label in [("신청","■ 심사 신청(진행중)"),("승인","■ 심사 승인"),("철회미승인","■ 철회·미승인")]:
        rows = kind_data["tables"][block]
        ws1.cell(r,1,f"{label}  ({len(rows)}건)").font = F(True,10,"1F3864")
        for c in range(1,len(cols1)+1): ws1.cell(r,c).fill = SUB
        r += 1
        for i,cn in enumerate(cols1,1):
            cell = ws1.cell(r,i,cn); cell.fill=HDR; cell.font=F(True,10,"FFFFFF"); cell.alignment=CEN; cell.border=BORD
        r += 1
        for row in rows:
            unit = row.get("단위","")
            vals = [row.get("일자",""), row.get("회사명",""),
                    "유가" if row.get("유가")=="Y" else "코스닥",
                    row.get("상장유형","") or row.get("심사결과",""),
                    f"FY{row['기준연도'][2:]}" if row.get("기준연도") else "",
                    row.get("매출액",""), row.get("순이익",""), row.get("자기자본",""),
                    row.get("업종",""), row.get("대표주관사",""), row.get("현황","")]
            for i,v in enumerate(vals,1):
                cell = ws1.cell(r,i,v); cell.font=F(); cell.border=BORD
                cell.alignment = RIG if i in (6,7,8) else CEN
            r += 1
        r += 1
    for col,w in zip("ABCDEFGHIJK",[11,16,7,12,8,12,11,11,26,14,16]):
        ws1.column_dimensions[col].width = w

    # ── 탭2 공모~상장 ──
    ws2 = wb.create_sheet("2_공모_상장")
    r = 1
    ws2.cell(r,1,"공모 ~ 상장 현황 (2026)").font = F(True,14); r += 1
    ws2.cell(r,1,'출처: DART 증권신고서(최신 정정본 "정정 후")·증권발행실적보고서·KIND 신규상장 · 공모금액=밴드 하단(확정 시 확정총액) · 단위 백만원').font = F(size=9,color="808080"); r += 2
    COLS=[("","","기업명",C_GRAY),
     ("신고서 제출+정정 시",C_BLUE,"수요예측 기간",None),("신고서 제출+정정 시",C_BLUE,"공모가밴드(원)",None),
     ("신고서 제출+정정 시",C_BLUE,"공모주식수량",None),("신고서 제출+정정 시",C_BLUE,"공모금액(백만원)",None),
     ("신고서 제출+정정 시",C_BLUE,"밸류에이션 멀티플",None),
     ("확정공모가 공시 시",C_GREEN,"확정공모가(원)",None),("확정공모가 공시 시",C_GREEN,"수요예측 경쟁률",None),
     ("확정공모가 공시 시",C_GREEN,"참여기관수",None),
     ("신고서 제출+정정 시",C_BLUE,"청약일정",None),
     ("청약완료(실적보고서)",C_ORANGE,"청약경쟁률",None),("청약완료(실적보고서)",C_ORANGE,"상장예정일",None),
     ("상장 완료(KIND)",C_RED,"상장일",None),
     ("신고서 제출+정정 시",C_BLUE,"대표주관사",None),("신고서 제출+정정 시",C_BLUE,"공동주관",None),
     ("신고서 제출+정정 시",C_BLUE,"인수수수료(백만원)",None),("신고서 제출+정정 시",C_BLUE,"수수료율",None),
     ("","","진행상태",C_GRAY)]
    hr1 = r
    for i,(t1,c1,t2,c2) in enumerate(COLS,1):
        cell = ws2.cell(hr1,i,t1 or ""); cell.fill = fill(c1 if t1 else C_GRAY)
        cell.font=F(True,8.5,"FFFFFF"); cell.alignment=CEN; cell.border=BORD
    hr2 = r+1
    for i,(t1,c1,t2,c2) in enumerate(COLS,1):
        cell = ws2.cell(hr2,i,t2); cell.fill = fill(c1 if c1 else (c2 or C_GRAY))
        cell.font=F(True,9.5,"FFFFFF"); cell.alignment=CEN; cell.border=BORD
    ws2.row_dimensions[hr1].hidden = True
    ws2.row_dimensions[hr1].height = 0.1
    r = hr2+1
    NCOL = len(COLS)

    def section(label, count):
        nonlocal r
        ws2.cell(r,1,f"■ {label}  ({count}건)").font = F(True,10,"1F3864")
        for c in range(1,NCOL+1): ws2.cell(r,c).fill = SUB
        r += 1
    def write_row(row):
        nonlocal r
        for i,v in enumerate(row,1):
            cell = ws2.cell(r,i,v); cell.border = BORD
            if v == V: cell.font=F(size=9,color="A6A6A6",italic=True); cell.alignment=CEN
            elif v == NA: cell.font=F(size=9,color="BFBFBF"); cell.alignment=CEN
            else:
                cell.font=F(); cell.alignment = RIG if i in NUMCOLS and v != "" else CEN
        ws2.row_dimensions[r].height = 30
        r += 1
    def section(label, count):  # noqa: F811  (섹션 헤더: 통일 높이)
        nonlocal r
        ws2.cell(r,1,f"■ {label}  ({count}건)").font = F(True,10,"1F3864")
        for c in range(1,NCOL+1): ws2.cell(r,c).fill = SUB
        ws2.row_dimensions[r].height = 22
        r += 1

    # 상장 완료: 시드 + 신규상장 페이지 병합 (상장일은 페이지 기준 갱신)
    _seed_list = _load_listing_seed()
    seed = {x["기업명"]: x for x in _seed_list}
    norm_index = {_norm_name(x["기업명"]): x["기업명"] for x in _seed_list}
    if listings:
        for l in listings:
            nm = l["회사명"]
            key = _norm_name(nm)
            if key in norm_index:                 # 정규화 일치 → 기존 시드 행 갱신(중복 방지)
                seed[norm_index[key]]["상장일"] = l["상장일"]
            elif nm in seed:
                seed[nm]["상장일"] = l["상장일"]
            else:                                  # 시드에 없는 신규 상장 → 추가
                seed[nm] = {"기업명": nm, "상장일": l["상장일"],
                            "진행상태": "상장완료" + ("(스팩합병)" if "합병" in l.get("상장유형","") else "")}
                norm_index[key] = nm
    FIELDS = ["기업명","수요예측기간","밴드","공모주식수","공모금액","멀티플","확정공모가",
              "수요예측경쟁률","참여기관수","청약일정","청약경쟁률","상장예정일","상장일",
              "대표주관","공동주관","인수수수료","수수료율","진행상태"]
    listed_rows = sorted(seed.values(),
        key=lambda x: x.get("상장일","") or "0", reverse=True)
    section("상장 완료 (2026)", len(listed_rows))
    for x in listed_rows:
        write_row([x.get(f, V if f not in ("공동주관","상장예정일") else "") for f in FIELDS])
    r += 1

    # 공모 진행: offering_master (상장일 확정된 건은 위 섹션으로 승격)
    prog = []
    for name, rec in offering.items():
        if name in seed and seed[name].get("상장일"):
            continue
        prog.append(rec)
    order = {"청약완료·상장대기":0,"발행조건확정":1,"정정":2,"최초":3,"미제출":9}
    prog.sort(key=lambda x: (order.get(x.get("상태",""),5), x.get("수요예측기간","") or "zz"))
    section("공모 진행 (예심 승인법인)", len(prog))
    for x in prog:
        write_row([x.get("회사명",""), x.get("수요예측기간",V), x.get("밴드",V),
                   x.get("공모주식수",V), x.get("공모금액",V), x.get("멀티플",V),
                   x.get("확정공모가",""), x.get("수요예측경쟁률",""), x.get("참여기관수",""),
                   x.get("청약일정",V), x.get("청약경쟁률",""), x.get("상장예정일",""),
                   "", x.get("대표주관",V), x.get("공동주관",""),
                   x.get("인수대가",V), x.get("수수료율",V),
                   {"미제출":"신고서 대기"}.get(x.get("상태"), x.get("상태",""))])
    r += 1
    for n in ['※ 상장 완료 = KIND 신규상장 페이지 기준(스팩합병 포함) · 스팩 공모상장은 밴드·밸류에이션 등 해당없음(공모가 2,000원 고정).',
              '※ 기재정정 신고서 내 정정 전·후 병존 → 자동수집은 "정정 후" 값만 취함.',
              '※ 청약경쟁률=일반투자자 청약수량÷최종배정수량 · "MM-DD~"는 종료일 자동수집 예정.']:
        ws2.cell(r,1,n).font = F(size=8,color="808080"); r += 1
    ws2.freeze_panes = f"B{hr2+1}"
    for i,w in enumerate([15,11,15,11,12,13,10,9,8,11,9,11,11,7,7,10,7,13],1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[hr2].height = 30

    # ── 탭3 분석 ──
    ws3 = wb.create_sheet("3_분석")
    r = 1
    ws3.cell(r,1,f"IPO 시장 분석 ({kind_data['year']})").font = F(True,14); r += 1
    ws3.cell(r,1,"승인율=결과확정 시점 기준 · 상장=KIND 신규상장 페이지 기준(스팩·합병 포함)").font = F(size=9,color="808080"); r += 2

    agg = defaultdict(lambda: defaultdict(int))
    # 승인 = 예심 통과(상장완료·미상장·후속철회 모두 포함) / 철회 = 예심 단계 반려
    APPROVED_ST = {"심사승인", "상장승인", "공모철회", "상장철회"}
    REJECTED_ST = {"심사철회", "심사미승인"}
    for rec in kind_master["records"].values():
        res = (rec.get("심사결과") or "").replace(" ", "")
        ty = (rec.get("상장유형") or "").replace(" ", "")
        cd = (rec.get("청구일") or "")[:7]
        rd = ((rec.get("결과확정일") or rec.get("청구일")) or "")[:7]
        if cd.startswith(kind_data["year"]): agg[cd]["청구"] += 1
        if rd.startswith(kind_data["year"]) and ty != "재상장":  # 재상장 제외
            if res in APPROVED_ST: agg[rd]["승인"] += 1
            elif res in REJECTED_ST: agg[rd]["철회"] += 1
    for l in (listings or []):
        m = (l.get("상장일") or "")[:7]
        if m.startswith(kind_data["year"]): agg[m]["상장"] += 1
    if not listings:
        for x in _load_listing_seed():
            m = (x.get("상장일") or "")[:7]
            if m.startswith(kind_data["year"]): agg[m]["상장"] += 1

    ws3.cell(r,1,"① 월별 예심·상장 현황 및 승인율").font = F(True,10,"1F3864")
    for c in range(1,7): ws3.cell(r,c).fill = SUB
    r += 1
    for i,cn in enumerate(["월","청구","승인","철회·미승인","상장","예심 승인율"],1):
        cell = ws3.cell(r,i,cn); cell.fill=HDR; cell.font=F(True,10,"FFFFFF"); cell.alignment=CEN; cell.border=BORD
    ds = r+1; r += 1
    for m in sorted(agg):
        a = agg[m]
        ws3.cell(r,1,m); ws3.cell(r,2,a["청구"]); ws3.cell(r,3,a["승인"])
        ws3.cell(r,4,a["철회"]); ws3.cell(r,5,a["상장"])
        ws3.cell(r,6,f'=IF((C{r}+D{r})=0,"-",C{r}/(C{r}+D{r}))')
        for c in range(1,7):
            cell = ws3.cell(r,c); cell.font=F(); cell.border=BORD; cell.alignment=CEN
            if c == 6: cell.number_format = "0.0%"
        r += 1
    ws3.cell(r,1,"합계").font=F(True); ws3.cell(r,1).alignment=CEN
    for c,col in zip(range(2,6),"BCDE"):
        cell = ws3.cell(r,c,f"=SUM({col}{ds}:{col}{r-1})"); cell.font=F(True); cell.alignment=CEN
    cell = ws3.cell(r,6,f'=IF((C{r}+D{r})=0,"-",C{r}/(C{r}+D{r}))')
    cell.font=F(True); cell.number_format="0.0%"; cell.alignment=CEN
    for c in range(1,7):
        ws3.cell(r,c).fill = fill("FCE4D6"); ws3.cell(r,c).border = BORD
    r += 2

    ws3.cell(r,1,"② 주관사별 점유율 (공모 확보분)").font = F(True,10,"1F3864")
    for c in range(1,4): ws3.cell(r,c).fill = SUB
    r += 1
    for i,cn in enumerate(["대표주관사","건수","공모금액(백만원)"],1):
        cell = ws3.cell(r,i,cn); cell.fill=HDR; cell.font=F(True,10,"FFFFFF"); cell.alignment=CEN; cell.border=BORD
    r += 1
    share = defaultdict(lambda: [0,0.0])
    src = list(seed.values()) + list(offering.values())
    for x in src:
        lead = (x.get("대표주관") or "").strip()
        if not lead or lead in (V, NA): continue
        amt = str(x.get("공모금액","")).replace(",","").split(" ")[0].split("(")[0]
        try: a = float(amt)
        except Exception: a = 0.0
        share[lead][0] += 1; share[lead][1] += a
    for lead,(n,a) in sorted(share.items(), key=lambda x:-x[1][1]):
        ws3.cell(r,1,lead); ws3.cell(r,2,n); ws3.cell(r,3,f"{a:,.0f}" if a else "")
        for c in range(1,4):
            cell = ws3.cell(r,c); cell.font=F(); cell.border=BORD; cell.alignment=CEN
        r += 1
    for col,w in zip("ABCDEF",[12,10,16,10,10,12]): ws3.column_dimensions[col].width = w

    wb.save(XLSX_OUT)
    print(f"[EXCEL] {os.path.basename(XLSX_OUT)} 생성 완료 "
          f"(상장 {len(listed_rows)} · 진행 {len(prog)})")

    # 웹 대시보드용 계산 결과 반환 (엑셀 탭2·탭3와 동일 값 재사용)
    return {
        "listed": listed_rows,
        "prog": prog,
        "monthly": [{"월": m, "청구": agg[m]["청구"], "승인": agg[m]["승인"],
                     "철회": agg[m]["철회"], "상장": agg[m]["상장"]} for m in sorted(agg)],
        "share": [{"대표주관사": k, "건수": v[0], "공모금액": v[1]}
                  for k, v in sorted(share.items(), key=lambda x: -x[1][1])],
    }


# ════════════════════════════════════════════════════════════════════
# [4-b] 웹 대시보드 (index.html) — 예심 / 공모·상장 / 분석 3탭 (기간필터·엑셀 다운로드)
#   ※ 데이터는 build_excel 이 계산한 값 그대로 재사용(엑셀과 동일).
#      수집·발췌 로직은 건드리지 않음.
# ════════════════════════════════════════════════════════════════════
DASH_TEMPLATE = """<!DOCTYPE html>
<html lang="ko"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>IPO 시장 동향</title>
<style>
  :root{--navy:#1F3864;--sub:#D9E1F2;--line:#D6DCE5;--muted:#6B7280;--bg:#F4F6FA;--red:#9A3B3B;}
  *{box-sizing:border-box;}
  body{margin:0;background:var(--bg);color:#1a1a1a;
       font-family:"맑은 고딕","Malgun Gothic",system-ui,sans-serif;font-size:14px;line-height:1.5;}
  .wrap{max-width:1180px;margin:0 auto;padding:22px 18px 60px;}
  h1{font-size:22px;margin:0 0 4px;color:var(--navy);}
  .asof{color:var(--muted);font-size:12.5px;margin-bottom:16px;}
  .tabs{display:flex;gap:6px;border-bottom:2px solid var(--navy);margin-bottom:2px;flex-wrap:wrap;}
  .tabs button{border:1px solid var(--line);border-bottom:none;background:#fff;color:var(--muted);
      padding:10px 18px;font-size:14px;font-weight:600;cursor:pointer;border-radius:8px 8px 0 0;}
  .tabs button.active{background:var(--navy);color:#fff;border-color:var(--navy);}
  .pane{display:none;background:#fff;border:1px solid var(--line);border-top:none;
        padding:18px 16px 24px;border-radius:0 0 10px 10px;}
  .pane.active{display:block;}
  .sec{font-weight:700;color:var(--navy);margin:18px 0 8px;font-size:15px;}
  .sec:first-child{margin-top:4px;}
  .cnt{color:var(--muted);font-weight:600;font-size:13px;}
  .note{color:var(--muted);font-size:12px;margin-top:6px;}
  .search{margin:2px 0 10px;}
  .search input{width:260px;max-width:70%;padding:8px 12px;border:1px solid var(--line);
      border-radius:8px;font-size:14px;}
  .tblwrap{overflow-x:auto;border:1px solid var(--line);border-radius:8px;}
  table{border-collapse:collapse;width:100%;font-size:13px;white-space:nowrap;}
  thead th{background:var(--navy);color:#fff;font-weight:600;padding:9px 10px;text-align:center;
      position:sticky;top:0;}
  tbody td{padding:8px 10px;border-top:1px solid #EEF1F6;text-align:center;}
  tbody tr:nth-child(even){background:#FafBfD;}
  tbody tr:hover{background:#EEF3FB;}
  td.num,th.num{text-align:right;font-variant-numeric:tabular-nums;}
  td.name{text-align:left;font-weight:600;}
  .dash{color:#C7CDD6;}
  .empty{color:var(--muted);padding:16px;}
  .totrow td{font-weight:700;background:#FCE4D6 !important;}
  .ctrl{display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin:2px 0 14px;
        padding:10px 12px;background:#F7F9FC;border:1px solid var(--line);border-radius:9px;}
  .ctrl label{font-size:12.5px;color:var(--muted);font-weight:700;}
  .ctrl .ipt{padding:7px 11px;border:1px solid var(--line);border-radius:7px;font-size:13.5px;width:190px;}
  .ctrl .dt{padding:6px 8px;border:1px solid var(--line);border-radius:7px;font-size:13px;}
  .preset{border:1px solid var(--line);background:#fff;color:var(--muted);
          padding:6px 10px;border-radius:6px;font-size:12.5px;cursor:pointer;}
  .preset:hover{background:var(--sub);}
  .btn-dl{margin-left:auto;border:1px solid var(--navy);background:var(--navy);color:#fff;
          padding:8px 15px;border-radius:7px;font-size:13px;font-weight:700;cursor:pointer;}
  .btn-dl:hover{opacity:.9;}
  .seg{display:inline-flex;border:1px solid var(--line);border-radius:7px;overflow:hidden;}
  .seg .segbtn{border:none;background:#fff;color:var(--muted);padding:8px 16px;
               font-size:13px;font-weight:700;cursor:pointer;}
  .seg .segbtn.active{background:var(--navy);color:#fff;}
</style>
<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
</head>
<body><div class="wrap">
  <h1>IPO 시장 동향</h1>
  <div class="asof" id="asof"></div>
  <div class="tabs">
    <button data-t="t1" class="active">예심</button>
    <button data-t="t2">공모 · 상장</button>
    <button data-t="t3">분석</button>
  </div>
  <div class="pane active" id="t1">
    <div class="ctrl">
      <input id="q" class="ipt" placeholder="회사명 검색…" autocomplete="off">
      <label>기간</label>
      <input type="date" id="t1from" class="dt"> ~ <input type="date" id="t1to" class="dt">
      <button class="preset" data-tab="t1" data-p="all">전체</button>
      <button class="preset" data-tab="t1" data-p="year">올해</button>
      <button class="preset" data-tab="t1" data-p="q">최근3개월</button>
      <button class="btn-dl" id="dl1">⬇ 엑셀</button>
    </div>
    <div id="t1body"></div>
  </div>
  <div class="pane" id="t2">
    <div class="ctrl">
      <div class="seg">
        <button class="segbtn active" data-v="listed">상장 완료</button>
        <button class="segbtn" data-v="prog">공모 진행</button>
      </div>
      <span id="t2date">
        <label>상장일</label>
        <input type="date" id="t2from" class="dt"> ~ <input type="date" id="t2to" class="dt">
        <button class="preset" data-tab="t2" data-p="all">전체</button>
        <button class="preset" data-tab="t2" data-p="year">올해</button>
        <button class="preset" data-tab="t2" data-p="q">최근3개월</button>
      </span>
      <button class="btn-dl" id="dl2">⬇ 엑셀</button>
    </div>
    <div id="t2body"></div>
  </div>
  <div class="pane" id="t3">
    <div class="ctrl">
      <label>기간(월)</label>
      <input type="date" id="t3from" class="dt"> ~ <input type="date" id="t3to" class="dt">
      <button class="preset" data-tab="t3" data-p="all">전체</button>
      <button class="preset" data-tab="t3" data-p="year">올해</button>
      <button class="preset" data-tab="t3" data-p="q">최근3개월</button>
      <button class="btn-dl" id="dl3">⬇ 엑셀</button>
    </div>
    <div id="t3body"></div>
  </div>
</div>
<script>
const D = __DATA__;
function isoAsof(){ var a=String(D.asof||''); return /^\\d{8}$/.test(a) ? a.slice(0,4)+'-'+a.slice(4,6)+'-'+a.slice(6,8) : a.slice(0,10); }
document.getElementById('asof').textContent =
  '기준일 ' + isoAsof() + ' · ' + D.year + '년 · 단위 백만원 · 출처 KIND·DART';

let T2VIEW = 'listed';
let EXPORT = {t1:null, t2:null, t3:null};
function val(id){ var e=document.getElementById(id); return e?(e.value||''):''; }
function inRange(d, from, to){
  if(!from && !to) return true;
  if(!d) return false;
  var x = String(d).slice(0,10);
  if(from && x < from) return false;
  if(to && x > to) return false;
  return true;
}
function monthInRange(mo, from, to){
  if(!from && !to) return true;
  if(!mo) return false;
  var f = from?from.slice(0,7):'', t = to?to.slice(0,7):'';
  if(f && mo < f) return false;
  if(t && mo > t) return false;
  return true;
}
function stamp(from,to){ return (!from && !to) ? '전체' : (from||'')+'~'+(to||''); }
function download(tabKey){
  var p = EXPORT[tabKey]; if(!p || typeof XLSX==='undefined') return;
  var wb = XLSX.utils.book_new();
  p.sheets.forEach(function(s){
    var aoa = [s.cols.map(function(c){return c[1];})];
    s.rows.forEach(function(r){ aoa.push(s.cols.map(function(c){ var v=r[c[0]]; return (v===undefined||v===null)?'':v; })); });
    XLSX.utils.book_append_sheet(wb, XLSX.utils.aoa_to_sheet(aoa), (s.name||'Sheet').slice(0,28));
  });
  XLSX.writeFile(wb, p.fname||'export.xlsx');
}
function applyPreset(tab,p){
  var today = isoAsof(), from='', to='';
  if(p==='year'){ from = D.year+'-01-01'; to = today; }
  else if(p==='q'){ var d=new Date(today); d.setMonth(d.getMonth()-3); from=d.toISOString().slice(0,10); to=today; }
  document.getElementById(tab+'from').value = from;
  document.getElementById(tab+'to').value = to;
  rerender(tab);
}
function rerender(tab){ if(tab==='t1') renderT1(); else if(tab==='t2') renderT2(); else renderT3(); }

function esc(v){ return (v===null||v===undefined)?'':String(v); }
function tbl(rows, cols, opts){
  opts = opts || {};
  let h = '<div class="tblwrap"><table><thead><tr>';
  h += cols.map(c=>`<th class="${c[2]?'num':''}">${c[1]}</th>`).join('');
  h += '</tr></thead><tbody>';
  if(!rows.length){
    h += `<tr><td class="empty" colspan="${cols.length}">해당 없음</td></tr>`;
  } else {
    for(const r of rows){
      const tot = opts.totalKey && r[opts.totalKey];
      h += tot ? '<tr class="totrow">' : '<tr>';
      h += cols.map((c,i)=>{
        const v = esc(r[c[0]]);
        const cls = (c[2]?'num':'') + (i===0 && !c[2] ? ' name':'');
        return `<td class="${cls}">${v===''?'<span class="dash">–</span>':v}</td>`;
      }).join('');
      h += '</tr>';
    }
  }
  return h + '</tbody></table></div>';
}

/* ── 탭1 예심 ── */
const SCR = [["일자","일자"],["회사명","회사명"],["시장","시장"],["상장유형","상장유형"],
  ["현황","현황"],["기준연도","기준연도"],["매출액","매출액(백만원)",1],["순이익","순이익(백만원)",1],
  ["자기자본","자기자본(백만원)",1],["업종","업종"],["대표주관사","대표주관사"]];
function renderT1(){
  const q = (val('q')).trim(), from = val('t1from'), to = val('t1to');
  const blocks = [["신청","심사 신청(진행중)"],["승인","심사 승인"],["철회미승인","철회·미승인"]];
  let h = ''; const sheets = [];
  for(const [k,label] of blocks){
    let rows = (D.screening[k] || []).filter(r => inRange(r['일자'], from, to));
    if(q) rows = rows.filter(r => (r['회사명']||'').includes(q));
    h += `<div class="sec">${label} <span class="cnt">(${rows.length}건)</span></div>`;
    h += tbl(rows, SCR);
    sheets.push({name: label.replace(/[^가-힣A-Za-z0-9]/g,''), cols: SCR, rows: rows});
  }
  document.getElementById('t1body').innerHTML = h;
  EXPORT.t1 = {sheets: sheets, fname: '예심_'+stamp(from,to)+'.xlsx'};
}

/* ── 탭2 공모·상장 ── */
const LISTED = [["기업명","기업명"],["상장일","상장일"],["밴드","공모가밴드(원)"],
  ["공모주식수","공모주식수(주)",1],["공모금액","공모금액(백만원)",1],
  ["수요예측경쟁률","수요예측경쟁률"],["참여기관수","수요예측 참여기관수",1],
  ["멀티플","밸류 멀티플"],["확정공모가","확정공모가(원)"],
  ["청약기간","청약기간"],["청약경쟁률","청약경쟁률"],["대표주관","대표주관"],["진행상태","상태"]];
const PROG = [["회사명","기업명"],["수요예측기간","수요예측기간"],["밴드","공모가밴드(원)"],
  ["공모주식수","공모주식수(주)",1],["공모금액","공모금액(백만원)",1],["멀티플","밸류 멀티플"],
  ["확정공모가","확정공모가(원)"],["수요예측경쟁률","수요예측경쟁률"],["참여기관수","수요예측 참여기관수",1],
  ["청약기간","청약기간"],["청약경쟁률","청약경쟁률"],["상장예정일","상장예정일"],["대표주관","대표주관"],["상태","상태"]];
function renderT2(){
  const dc = document.getElementById('t2date');
  if(dc) dc.style.display = (T2VIEW==='listed') ? '' : 'none';
  let h = '';
  if(T2VIEW === 'listed'){
    const from = val('t2from'), to = val('t2to');
    const rows = D.listed.filter(r => inRange(r['상장일'], from, to));
    h += `<div class="sec">상장 완료 (${D.year}) <span class="cnt">(${rows.length}건)</span></div>`;
    h += tbl(rows, LISTED);
    h += '<div class="note">※ 상장 완료 = KIND 신규상장 페이지 기준(스팩합병 포함) · '
       + '공모금액 = 밴드 하단(확정 시 확정총액) · 상장일 기준 기간 조회.</div>';
    EXPORT.t2 = {sheets:[{name:'상장완료', cols:LISTED, rows:rows}], fname:'상장완료_'+stamp(from,to)+'.xlsx'};
  } else {
    const rows = D.prog;
    h += `<div class="sec">공모 진행 (예심 승인법인) <span class="cnt">(${rows.length}건)</span></div>`;
    h += tbl(rows, PROG);
    h += '<div class="note">※ 공모 진행 = 예심 승인 후 공모 절차 진행 중(미상장). 기간 필터 미적용.</div>';
    EXPORT.t2 = {sheets:[{name:'공모진행', cols:PROG, rows:rows}], fname:'공모진행.xlsx'};
  }
  document.getElementById('t2body').innerHTML = h;
}

/* ── 탭3 분석 (일반 + 비밀) ── */
const MONTH = [["월","월"],["청구","청구",1],["승인","승인",1],
  ["철회","철회·미승인",1],["상장","상장",1],["승인율","예심 승인율"]];
const KV = [["항목","항목"],["값","값"]];
const YO = [["월","월"],["기업수","기업수",1],["경쟁률평균","수요예측 경쟁률(평균)"],["참여평균","참여기관수(평균)",1]];
const CY = [["월","월"],["기업수","청약 기업수",1],["경쟁률평균","청약 경쟁률(평균)"]];
const RANK_AMT = [["순위","순위",1],["증권사","증권사"],["건수","건수",1],["인수금액","인수금액(억)",1],["점유율","점유율"]];
const RANK_FEE = [["순위","순위",1],["증권사","증권사"],["건수","건수",1],["인수수수료","인수수수료(억)",1]];
const RANK_CNT = [["순위","순위",1],["증권사","증권사"],["건수","상장건수",1]];
function renderT3(){
  const from = val('t3from'), to = val('t3to');
  const src = D.monthly.filter(m => monthInRange(m['월'], from, to));
  const tot = {청구:0, 승인:0, 철회:0, 상장:0};
  const rows = src.map(function(m){
    tot.청구+=m.청구; tot.승인+=m.승인; tot.철회+=m.철회; tot.상장+=m.상장;
    var den = m.승인 + m.철회;
    return {월:m.월, 청구:m.청구, 승인:m.승인, 철회:m.철회, 상장:m.상장,
            승인율: den ? (m.승인/den*100).toFixed(1)+'%' : '-'};
  });
  var tden = tot.승인 + tot.철회;
  const disp = rows.concat([{월:'합계', 청구:tot.청구, 승인:tot.승인, 철회:tot.철회, 상장:tot.상장,
            승인율: tden ? (tot.승인/tden*100).toFixed(1)+'%' : '-', 합계:true}]);
  let h = '<div class="sec">① 월별 예심 · 상장 현황 및 승인율</div>';
  h += tbl(disp, MONTH, {totalKey:"합계"});
  h += '<div class="sec">② 수요예측 현황 <span class="cnt">(상장 월별 평균 · 수집분 기준)</span></div>';
  h += tbl(D.yoyeuk||[], YO, {totalKey:"합계"});
  h += '<div class="sec">③ 청약 현황 <span class="cnt">(상장 월별 평균)</span></div>';
  h += tbl(D.cheongyak||[], CY, {totalKey:"합계"});
  h += '<div class="sec">④ 추가 분석 데이터 <span class="cnt">(비밀번호 필요)</span></div>';
  h += '<div class="ctrl"><label>비밀번호</label><input type="password" id="t3pw" class="dt" style="width:130px" autocomplete="off"><button class="preset" id="t3unlock">확인</button></div>';
  h += '<div id="t3secret"></div>';
  document.getElementById('t3body').innerHTML = h;
  var ub=document.getElementById('t3unlock'); if(ub) ub.onclick=unlockSecret;
  var pw=document.getElementById('t3pw'); if(pw) pw.addEventListener('keydown',function(e){ if(e.key==='Enter') unlockSecret(); });
  EXPORT.t3 = {sheets:[{name:'월별현황', cols:MONTH, rows:disp}], fname:'분석_'+stamp(from,to)+'.xlsx'};
}
function unlockSecret(){
  var box=document.getElementById('t3secret');
  var pw=(document.getElementById('t3pw')||{}).value||'';
  if(pw!=='1111*'){ box.innerHTML='<div class="note" style="color:var(--red)">비밀번호가 올바르지 않습니다.</div>'; return; }
  var L=(D.league||[]).slice();
  function rank(key){ return L.slice().sort(function(a,b){return (b[key]||0)-(a[key]||0);}).map(function(r,i){ return Object.assign({순위:i+1}, r); }); }
  var totAmt=L.reduce(function(s,r){return s+(r.인수금액||0);},0);
  var amtRows=rank('인수금액').map(function(r){ return Object.assign({}, r, {점유율:(totAmt?(r.인수금액/totAmt*100):0).toFixed(1)+'%'}); });
  var h='<div class="note">FY26 · 납입일 기준 누적 · 리그 검증 · 단위 억원 · 스팩·리츠 포함</div>';
  h+='<div class="sec">1. 공모금액(인수금액) 순위</div>'+tbl(amtRows, RANK_AMT);
  h+='<div class="sec">2. 인수수수료 순위</div>'+tbl(rank('인수수수료'), RANK_FEE);
  h+='<div class="sec">3. 상장 건수 순위</div>'+tbl(rank('건수'), RANK_CNT);
  h+='<button class="btn-dl" id="dlLedger" style="margin-top:14px;margin-left:0">⬇ raw 원장 다운로드</button>';
  box.innerHTML=h;
  var dl=document.getElementById('dlLedger'); if(dl) dl.onclick=downloadLedger;
}
function downloadLedger(){
  if(typeof XLSX==='undefined') return;
  var rows=[["업체","상장일","시장","증권사","역할","인수금액(억)","인수수수료(억)","청약수수료(억)","전체수수료(억)"]];
  (D.ledger||[]).forEach(function(d){ (d.인수인||[]).forEach(function(u){ rows.push([d.업체,d.상장일,d.시장,u.증권사,u.역할,u.인수금액,u.인수수수료,u.청약수수료,u.전체수수료]); }); });
  var ag=[["증권사","건수","인수금액(억)","인수수수료(억)","청약수수료(억)","전체수수료(억)"]];
  (D.league||[]).forEach(function(r){ ag.push([r.증권사,r.건수,r.인수금액,r.인수수수료,r.청약수수료,r.전체수수료]); });
  var wb=XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb, XLSX.utils.aoa_to_sheet(rows), "딜원장");
  XLSX.utils.book_append_sheet(wb, XLSX.utils.aoa_to_sheet(ag), "주관사집계");
  XLSX.writeFile(wb, "IPO_트랙레코드_원장.xlsx");
}

document.querySelectorAll('.tabs button').forEach(b=>{
  b.onclick = ()=>{
    document.querySelectorAll('.tabs button').forEach(x=>x.classList.remove('active'));
    document.querySelectorAll('.pane').forEach(x=>x.classList.remove('active'));
    b.classList.add('active');
    document.getElementById(b.dataset.t).classList.add('active');
  };
});
document.querySelectorAll('.seg .segbtn').forEach(b=>{
  b.onclick = ()=>{
    document.querySelectorAll('.seg .segbtn').forEach(x=>x.classList.remove('active'));
    b.classList.add('active'); T2VIEW = b.dataset.v; renderT2();
  };
});
document.querySelectorAll('.preset').forEach(b=>{ b.onclick = ()=>applyPreset(b.dataset.tab, b.dataset.p); });
document.getElementById('q').addEventListener('input', renderT1);
['t1from','t1to'].forEach(id=>document.getElementById(id).addEventListener('change', renderT1));
['t2from','t2to'].forEach(id=>document.getElementById(id).addEventListener('change', renderT2));
['t3from','t3to'].forEach(id=>document.getElementById(id).addEventListener('change', renderT3));
document.getElementById('dl1').onclick = ()=>download('t1');
document.getElementById('dl2').onclick = ()=>download('t2');
document.getElementById('dl3').onclick = ()=>download('t3');
renderT1(); renderT2(); renderT3();
</script></body></html>
"""

def build_dashboard(kind_data, web):
    """엑셀과 동일한 데이터로 3탭 웹페이지(index.html)를 생성."""
    def scr(rows):
        out = []
        for x in rows:
            out.append({
                "일자": x.get("일자",""), "회사명": x.get("회사명",""),
                "시장": "유가" if x.get("유가") == "Y" else "코스닥",
                "상장유형": x.get("상장유형","") or x.get("심사결과",""),
                "현황": x.get("현황",""),
                "기준연도": x.get("기준연도",""), "매출액": x.get("매출액",""),
                "순이익": x.get("순이익",""), "자기자본": x.get("자기자본",""),
                "업종": x.get("업종",""), "대표주관사": x.get("대표주관사","")})
        return out

    def _comma(v):
        s = str(v).replace(",", "").strip()
        return f"{int(s):,}" if s.isdigit() else (str(v) if v not in (None, "") else "")
    def _fmt_sub(s):
        m = re.match(r"\s*(\d{4})-(\d{1,2})-(\d{1,2})\s*~\s*(?:(\d{1,2})-)?(\d{1,2})", str(s or ""))
        if not m: return str(s or "")
        y, mo, d1, mo2, d2 = m.groups()
        mo, d1, d2 = int(mo), int(d1), int(d2)
        if mo2 and int(mo2) != mo:
            return f"{y}-{mo:02d}-{d1:02d}~{int(mo2):02d}-{d2:02d}"   # 월 바뀌면 MM-DD~MM-DD
        return f"{y}-{mo:02d}-{d1:02d}~{d2:02d}"                        # 같은 월이면 MM-DD~DD

    listed = [{
        "기업명": x.get("기업명",""), "상장일": x.get("상장일",""),
        "밴드": x.get("밴드",""), "공모주식수": x.get("공모주식수",""),
        "공모금액": x.get("공모금액",""), "멀티플": x.get("멀티플",""),
        "확정공모가": _comma(x.get("확정공모가","")), "참여기관수": x.get("참여기관수",""),
        "수요예측경쟁률": x.get("수요예측경쟁률",""),
        "청약기간": _fmt_sub(x.get("청약기간") or x.get("청약일정","")),
        "청약경쟁률": x.get("청약경쟁률",""), "대표주관": x.get("대표주관",""),
        "진행상태": x.get("진행상태","")} for x in web["listed"]]

    prog = [{
        "회사명": x.get("회사명",""), "수요예측기간": _fmt_sub(x.get("수요예측기간","")),
        "밴드": x.get("밴드",""), "공모주식수": x.get("공모주식수",""),
        "공모금액": x.get("공모금액",""), "멀티플": x.get("멀티플",""),
        "확정공모가": _comma(x.get("확정공모가","")), "참여기관수": x.get("참여기관수",""),
        "수요예측경쟁률": x.get("수요예측경쟁률",""),
        "청약기간": _fmt_sub(x.get("청약일정","")),
        "청약경쟁률": x.get("청약경쟁률",""), "상장예정일": x.get("상장예정일",""),
        "대표주관": x.get("대표주관",""),
        "상태": {"미제출":"신고서 대기"}.get(x.get("상태"), x.get("상태",""))} for x in web["prog"]]

    # 월별은 원자료 그대로 전달 (승인율·합계는 웹에서 기간필터에 맞춰 재계산)
    monthly = [{"월": m["월"], "청구": m["청구"], "승인": m["승인"],
                "철회": m["철회"], "상장": m["상장"]} for m in web["monthly"]]

    share = [{"대표주관사": s["대표주관사"], "건수": s["건수"],
              "공모금액": f'{s["공모금액"]:,.0f}' if s["공모금액"] else ""} for s in web["share"]]

    # 수요예측·청약 현황 (상장 월별 평균 + 누적/합계 · 수집분 기준)
    def _f(v):
        try: return float(str(v).replace(",", "").replace(":1", ""))
        except Exception: return None
    def _avg(a): return (sum(a) / len(a)) if a else None
    yo_m, cy_m = {}, {}
    for x in listed:
        d = (x.get("상장일") or "")
        mo = d[:7] if len(d) >= 7 else ""
        if not mo: continue
        v = _f(x.get("수요예측경쟁률")); iv = _f(x.get("참여기관수"))
        if v is not None:
            yo_m.setdefault(mo, {"c": [], "i": []})
            yo_m[mo]["c"].append(v)
            if iv is not None: yo_m[mo]["i"].append(iv)
        cv = _f(x.get("청약경쟁률"))
        if cv is not None:
            cy_m.setdefault(mo, []).append(cv)
    yoyeuk, allc, alli = [], [], []
    for mo in sorted(yo_m):
        c, i = yo_m[mo]["c"], yo_m[mo]["i"]; allc += c; alli += i
        yoyeuk.append({"월": mo, "기업수": len(c),
            "경쟁률평균": f"{_avg(c):,.1f}:1" if c else "-",
            "참여평균": f"{_avg(i):,.0f}" if i else "-"})
    yoyeuk.append({"월": "누적", "기업수": len(allc),
        "경쟁률평균": f"{_avg(allc):,.1f}:1" if allc else "-",
        "참여평균": f"{_avg(alli):,.0f}" if alli else "-", "합계": True})
    cheongyak, allcy = [], []
    for mo in sorted(cy_m):
        v = cy_m[mo]; allcy += v
        cheongyak.append({"월": mo, "기업수": len(v),
            "경쟁률평균": f"{_avg(v):,.1f}:1" if v else "-"})
    cheongyak.append({"월": "합계", "기업수": len(allcy),
        "경쟁률평균": f"{_avg(allcy):,.1f}:1" if allcy else "-", "합계": True})

    # 주관사 트랙레코드 원장 (리그 검증 정본)
    league, ledger = [], []
    lp = os.path.join(BASE, "underwriter_ledger.json")
    if os.path.exists(lp):
        try:
            LJ = json.load(open(lp, encoding="utf-8"))
            league = LJ.get("주관사집계", []); ledger = LJ.get("딜원장", [])
        except Exception as e:
            print(f"  [경고] 원장 로드 실패: {e}")

    DATA = {"asof": kind_data["asof"], "year": kind_data["year"],
            "screening": {k: scr(v) for k, v in kind_data["tables"].items()},
            "listed": listed, "prog": prog, "monthly": monthly, "share": share,
            "yoyeuk": yoyeuk, "cheongyak": cheongyak, "league": league, "ledger": ledger}
    html = DASH_TEMPLATE.replace("__DATA__", json.dumps(DATA, ensure_ascii=False))
    with open(os.path.join(BASE, "index.html"), "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[WEB] index.html 3탭 생성 완료 "
          f"(예심 {sum(len(v) for v in kind_data['tables'].values())} · "
          f"상장 {len(listed)} · 진행 {len(prog)})")


# ════════════════════════════════════════════════════════════════════
# [5] 원장 CSV 내보내기 (분석·가공용 · 매 실행 시 최신으로 재생성)
# ════════════════════════════════════════════════════════════════════
def export_ledgers(kind_master, offering, listings):
    import csv
    # ── events.csv : 이벤트 원장 (일자·회사·이벤트·상세) ──
    ev = []
    for rec in kind_master["records"].values():
        nm = rec.get("회사명", ""); ty = rec.get("상장유형", "")
        if rec.get("청구일"):
            ev.append([rec["청구일"], nm, "예심청구", ty])
        res = (rec.get("심사결과") or "").replace(" ", "")
        rd = (rec.get("결과확정일") or "").strip()
        if rd and res and res != "청구서접수":
            label = {"심사승인": "예심승인", "심사철회": "심사철회",
                     "심사미승인": "심사미승인", "상장승인": "상장승인"}.get(res, res)
            ev.append([rd, nm, label, ty])
    for l in (listings or []):
        ev.append([l.get("상장일", ""), l.get("회사명", ""), "상장",
                   l.get("상장유형", "")])
    if not listings:
        for x in _load_listing_seed():
            if x.get("상장일", "").startswith("20"):
                ev.append([x["상장일"], x["기업명"], "상장", x.get("진행상태", "")])
    for name, rec in (offering or {}).items():
        rn = rec.get("최신접수번호") or ""
        if len(rn) >= 8:
            ev.append([f"{rn[:4]}-{rn[4:6]}-{rn[6:8]}", name,
                       f"신고서({rec.get('상태','')})",
                       rec.get("확정공모가", "") or rec.get("밴드", "")])
        pr = rec.get("실적보고서") or ""
        if len(pr) >= 8:
            ev.append([f"{pr[:4]}-{pr[4:6]}-{pr[6:8]}", name, "발행실적보고",
                       rec.get("청약경쟁률", "")])
    ev = sorted([e for e in ev if e[0]], key=lambda x: x[0])
    with open(os.path.join(BASE, "events.csv"), "w",
              newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f); w.writerow(["일자", "회사명", "이벤트", "상세"]); w.writerows(ev)

    # ── offerings.csv : 공모 원장 (기업 1행 플랫 테이블) ──
    FIELDS = ["기업명", "진행상태", "수요예측기간", "밴드", "공모주식수", "공모금액",
              "멀티플", "확정공모가", "수요예측경쟁률", "참여기관수", "청약일정",
              "청약경쟁률", "상장예정일", "상장일", "대표주관", "공동주관",
              "인수수수료", "수수료율"]
    seed = {x["기업명"]: dict(x) for x in _load_listing_seed()}
    if listings:
        for l in listings:
            nm = l["회사명"]
            seed.setdefault(nm, {"기업명": nm, "진행상태": "상장완료"})
            seed[nm]["상장일"] = l["상장일"]
    rows = list(seed.values())
    for name, rec in (offering or {}).items():
        if name in seed and seed[name].get("상장일", "").startswith("20"):
            continue
        d = dict(rec); d["기업명"] = name
        d["진행상태"] = rec.get("상태", ""); d["인수수수료"] = rec.get("인수대가", "")
        rows.append(d)
    with open(os.path.join(BASE, "offerings.csv"), "w",
              newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f); w.writerow(FIELDS)
        for x in rows: w.writerow([x.get(k, "") for k in FIELDS])
    print(f"[LEDGER] events.csv {len(ev)}행 · offerings.csv {len(rows)}행")


# ════════════════════════════════════════════════════════════════════
# 메인
# ════════════════════════════════════════════════════════════════════
def main():
    print("=" * 62)
    print(f" IPO 통합 일일 자동화  ({TODAY})")
    print("=" * 62)

    # [1] KIND 예심 + index.html
    kind_data, kind_master = collect_kind()

    # [2] KIND 신규상장 (상장일 확정 소스)
    sess = requests.Session()
    listings = fetch_new_listings(sess)
    if listings is not None:
        print(f"[KIND] 신규상장 페이지: {len(listings)}건")
    else:
        if os.path.exists(LISTING_JSON):
            listings = json.load(open(LISTING_JSON, encoding="utf-8"))
            print(f"[KIND] 신규상장 접속 실패 → 기존 정본 사용({len(listings)}건)")
        else:
            listings = []
            print("[KIND] 신규상장 접속 실패 → 시드(listing_seed.json)로 대체")

    # [2-b] 스팩합병 상장 병합 (신규상장 페이지에는 안 잡히므로 예심 master에서 보강)
    #   기준: 심사결과='상장승인' & 상장유형에 '스팩'+'합병' & 당해연도
    have = {l["회사명"] for l in listings}
    spac_added = 0
    for rec in kind_master["records"].values():
        ltype = (rec.get("상장유형") or "").replace(" ", "")
        res = (rec.get("심사결과") or "").replace(" ", "")
        rd = (rec.get("결과확정일") or "")
        if res == "상장승인" and "스팩" in ltype and "합병" in ltype \
                and rd.startswith(str(datetime.date.today().year)):
            nm = rec["회사명"]
            if nm not in have:
                listings.append({"회사명": nm, "상장일": rd,
                                 "상장유형": "스팩합병", "시장": "코스닥"})
                have.add(nm); spac_added += 1
    if spac_added:
        print(f"[KIND] 스팩합병 상장 병합: +{spac_added}건 "
              f"(합계 {len(listings)}건)")
    json.dump(listings, open(LISTING_JSON, "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)

    # [3] DART 공모 추적 (승인법인, 재상장 제외)
    approved = []
    for rec in kind_master["records"].values():
        if (rec.get("심사결과","").replace(" ","") == "심사승인"
                and "재상장" not in (rec.get("상장유형") or "")
                and (rec.get("결과확정일","") or "").startswith(kind_data["year"])):
            approved.append(rec["회사명"])
    if not DART_API_KEY:
        print("[DART] API 키 없음 → 공모 추적 생략 (환경변수 DART_API_KEY 설정 필요)")
        offering = json.load(open(OFFERING_JSON, encoding="utf-8")) \
                   if os.path.exists(OFFERING_JSON) else {}
    else:
        offering, ch = track_offerings(approved)
        print(f"[DART] 공모 추적: {len(offering)}건 · 변경 {len(ch)}건")
        for c in ch: print("   -", c)

    # [3-c] 신규 상장딜 신고서 자동발췌 → listing_seed.json · underwriter_ledger.json 갱신
    corp_map = dict(CORP_SEED)
    _cm = os.path.join(BASE, "corp_map.json")
    if os.path.exists(_cm):
        try: corp_map.update(json.load(open(_cm, encoding="utf-8")).get("corp_map", {}))
        except Exception: pass
    try:
        enrich_listed_from_dart(listings, corp_map)
    except Exception as e:
        print(f"[ENRICH] 실패(무시): {e}")

    # [4] 3탭 엑셀
    web = build_excel(kind_data, kind_master, offering, listings)

    # [4-b] 3탭 웹 대시보드 (index.html 재생성 · 엑셀과 동일 데이터)
    build_dashboard(kind_data, web)

    # [5] 원장 CSV (분석·가공용)
    export_ledgers(kind_master, offering, listings)
    print("\n[완료] IPO_analysis.xlsx / index.html(3탭) / events.csv / offerings.csv 갱신")

if __name__ == "__main__":
    main()
